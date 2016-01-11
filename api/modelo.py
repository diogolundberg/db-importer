""" Classes usadas para fazer os mapeamentos do banco de dados. Todos os
mapeamentos irão importar este módulo e herdarão de suas classes.
"""
from source import db
from api.sql import executar_sql
from api.importacao import distribuir_importacao, executar_importacao
from openpyxl import load_workbook


class Importavel(object):

    """ Classe que representa um recurso a ser importado, ele contém todo o
    necessário para que o importador saiba de onde tirar os dados, onde e como
    inserir.
    """

    def __init__(self, database_select=db.SASCWEB, database_insert=db.SASC):
        """ Método construtor. Está garantindo que o objeto terá um database de
        origem e destino, para a importação.

        Args:
            database_select (Config): Configurações do banco de dados origem.
            database_insert (Config): Configurações do banco de dados destino.

        Returns:
            Um objeto Importável.
        """
        self.database_select = database_select
        self.database_insert = database_insert
        self.tabela = None
        self.query = None
        self.orderby = None

    def count(self):
        """ Envolve a query deste importável com um count, para saber o número
        de registros

        Returns:
            Total de registros existentes deste importável no banco de origem.
        """
        cursor = db.cursor(self.database_select)
        count = cursor.execute_scalar(
            'SELECT count(*) FROM (' + self.query + ') as count')
        cursor.close()
        return count

    def select(self, offset, limit):
        """ Função utilizada internamente pelo módulo, que deve usar o cursor para
        selecionar no banco de dados a query do importável correspondente, com limit
        e offset.

        Args:
            cursor (Cursor): Objeto de um cursor, obtido através da conexão com o
                             banco de dados.
            offset (int): Este é o número do ponto de partida do select.
            limit (int): É o número de tuplas máximo que o select pode retornar.
            importavel (Importavel): Um objeto da classe Importavel.

        Returns:
            O cursor com o resultado da query executado no banco de origem do
            Importável.
        """
        cursor = db.cursor(self.database_select)
        if limit > 0:
            cursor.execute_query(self.query
                                 + ' ORDER BY '
                                 + self.orderby
                                 + '''
                                  OFFSET %s ROWS
                                  FETCH NEXT %s ROWS ONLY
                                  ''' % (offset, limit))
        else:
            cursor.execute_query(self.query)
        return cursor

    def importar(self,
                 fatia=None,
                 numero_threads=1,
                 tamanho_fila=10,
                 resolver_dependencias=False):
        """ Este método realiza a rotina de importação completa, utilizando o
        módulo de importação. Essa rotina de importação acontece nesta ordem:
            1 - As colunas temporárias de dependencias são criadas.
            2 - Os dados são lidos do banco de origem e inseridos no destino,
            fazendo uso das colunas para integridade.
            3 - Caso seja informado para resolver dependencias, a integridade
            será resolvida.

        Args:
            fatia (Array[int]): O trecho que será importado, exemplo: [1, 900]
            numero_threads (int): O número de threads paralelas.
            tamanho_fila (int): É a quantidade total de tarefas que o importador
                                será divido.
            resolver_dependencias (Boolean): Quando informado como True, o importador
                                             irá automaticamente realizar a
                                             integridade das tableas.
        """
        if fatia is None:
            fatia = [1, self.count()]

        self.executar_dependencias('add_colunas')

        print('\nImportando ' + str(self))
        if numero_threads == 1:
            executar_importacao(importavel=self, offset=fatia[0] - 1,
                                limit=fatia[1])
        else:
            distribuir_importacao(importavel=self,
                                  primeira_row=fatia[0],
                                  ultima_row=fatia[1],
                                  numero_threads=numero_threads,
                                  tamanho_fila=tamanho_fila)

        if resolver_dependencias:
            self.resolver_dependencias()

        print('\nImportação de ' + str(self) + ' concluída')

    def __str__(self):
        """ Sobreescrita do método de classe str, que é a forma que a o objeto
        é descrito como string.

        Returns:
            Nome da classe.
        """
        return str(self.__class__.__name__)

    def lista_dependencias(self):
        """ Método que lista todos os atributos declarados do tipo Dependencia

        Returns:
            Lista com todos os objetos dentro deste importável do tipo
            Dependencia.
        """
        dependencias = [dependencia for dependencia in vars(self).values()
                        if isinstance(dependencia, Dependencia)]
        return dependencias

    def resolver_dependencias(self):
        """ Método que lista todos os atributos declarados do tipo Dependencia

        Returns:
            Lista com todos os objetos dentro deste importável do tipo
            Dependencia.
        """
        self.executar_dependencias('update_fk')

    def dropar_colunas_temporarias(self):
        """ Método que lista todos os atributos declarados do tipo Dependencia

        Returns:
            Lista com todos os objetos dentro deste importável do tipo
            Dependencia.
        """
        self.executar_dependencias('drop_colunas')

    def executar_dependencias(self, funcao):
        """ Este método executa em todas as dependencias do importável a função
        com o nome fornecido por parâmetro.

        Args:
            funcao (String): Nome do método a ser executado pelas dependências.

        Yields:
            Executa em todas as dependências o método com o nome igual ao
            parâmetro função.
        """
        dependencias = self.lista_dependencias()
        if dependencias:
            executar_sql(self.database_insert, '\n'.join(
                [getattr(dependencia, funcao)() for dependencia in dependencias]))

    def dependencia(self, fk_, colunas, tabela_dependencia, **condicoes):
        """ Método que cria uma dependência para este importável.

        Args:
            fk_ (String): Nome da coluna com a FK
            colunas (Array[dict]): Lista com dicionarios fazendo o 'de para' das
                                   colunas da dependência.
            tabela_dependencia (String): Nome da tabela para qual a FK aponta.
            **condicoes (String): Condições de igualdade entre as colunas listadas.
                                  Não é obrigatório, caso nenhuma condição seja
                                  fornecida, a dependência considerará somente
                                  a primeira coluna.
        Returns:
            Um objeto Dependência.
        """
        return Dependencia(importavel=self,
                           fk_=fk_,
                           colunas=colunas,
                           tabela_dependencia=tabela_dependencia,
                           **condicoes)


class Dependencia(object):

    """ Classe que representa uma dependencia de um importável, contém ps métodos
    para gerar a integridade.
    """

    def __init__(self, importavel, fk_, colunas, tabela_dependencia, **condicoes):
        """ Método construtor. Garante que o objeto terá todos os atributos
        de uma dependência.

        Args:
            importavel (Importavel): O importável que possui esta dependência.
            fk_ (String): Nome da coluna com a FK
            colunas (Array[dict]): Lista com dicionarios fazendo o 'de para' das
                                   colunas da dependência.
            tabela_dependencia (String): Nome da tabela para qual a FK aponta.
            **condicoes (String): Condições de igualdade entre as colunas listadas.
                                  Não é obrigatório, caso nenhuma condição seja
                                  fornecida, a dependência considerará somente
                                  a primeira coluna.

        Returns:
            Um objeto Dependência.
        """
        self.importavel = importavel
        self.tabela = importavel.tabela
        self.tabela_dependencia = tabela_dependencia
        self.fk_ = fk_
        self.colunas = colunas
        self.condicoes = self.condicoes_integridade(**condicoes)

    def condicoes_integridade(self, **args):
        """ Esse método concatena a sentença para condição de integridade da
        dependência.

        Args:
            **args (Kwargs): É uma de parâmetros nomeados dinâmica.

        Returns:
            A sentença SQL com as condições para gerar a integridade da Dependência.
        """
        condicoes = []

        for coluna in self.colunas.keys():
            colunafk = '.' + coluna + '_' + self.fk_
            coluna = '.' + (args[coluna] if coluna in args else coluna)
            condicoes.append(self.tabela_dependencia + coluna +
                             ' = ' + self.tabela + colunafk)

        return '\nAND '.join(condicoes)

    def dados(self, row):
        """ Este método deve criar a parte da dependência inserida no mapeamentos.
        Isto representará a tradução da fk desta dependência para o recurso que
        está sendo inserido.
        Os métodos 'dados' são utilizados pela tarefa de importação para obter o
        'de para' dos recursos mapeados. Sempre que o recurso possuir dependencias,
        o mesmo método das dependencias será utilizado.

        Args:
            row (Dict): É um dicinário, com chaves de nomes de colunas e valores
                        dos campos selecionados. A row é usada como referência
                        a linha do resultado do select do importável, durante a
                        importação.

        Return:
            O retorno será sempra um mapa com os dados puros do 'de para' para a
            row referenciada no parâmetro, por exemplo:
                {
                    'fk_1':       valor_fk_1,
                    'fk_2':       valor_fk_2,
                    'coluna_1':   valor_1,
                    'coluna_2':   valor_2,
                    'coluna_3':   valor_3,
                    'coluna_4':   valor_4,
                    'coluna_5':   valor_5,
                }
        """
        dados = dict()
        for coluna_imp, coluna_select in self.colunas.items():
            dados.update(
                {
                    coluna_imp + '_' + self.fk_: coluna_select(row) if callable(coluna_select)
                                                 else str(row[coluna_select]).strip()
                })
        return dados

    def update_fk(self):
        """ Método que faz a sentença sql para realizar integridade desta
        dependencia.

        Returns:
            A sentença SQL com o comando realizar update em todas as FKs da
            dependências, gerando integridade.
        """
        return ("\nUPDATE " + self.tabela +
                " SET " + self.fk_ +
                " = (SELECT id FROM " + self.tabela_dependencia +
                " WHERE \n" + self.condicoes + ");")

    def drop_colunas(self):
        """ Método que faz a sentença em sql para apagar as colunas temporárias
        de integridade.

        Returns:
            A sentença SQL com o comando para dropar as colunas temporárias da
            dependência.
        """
        sql = ''
        for coluna in self.colunas.keys():
            sql += ("\nIF exists( "
                    " SELECT * FROM INFORMATION_SCHEMA.COLUMNS "
                    " WHERE COLUMN_NAME = '" + coluna + '_' + self.fk_ +
                    "' AND TABLE_NAME = '" +
                    self.tabela + "') begin"
                    "\nALTER TABLE " + self.tabela
                    + " DROP COLUMN " + coluna + '_' + self.fk_ + ";"
                    "\nend;")
        return sql

    def add_colunas(self):
        """ Método que faz a sentença em sql para criar as colunas temporárias
        desta dependência.

        Returns:
            A sentença SQL com o comando para criar as colunas temporárias da
            dependência.
        """
        sql = ''
        for coluna in self.colunas.keys():
            sql += ("\nIF NOT exists("
                    " SELECT * FROM INFORMATION_SCHEMA.COLUMNS"
                    " WHERE COLUMN_NAME = '" + coluna + '_' + self.fk_ +
                    "' AND TABLE_NAME = '" +
                    self.tabela + "') begin"
                    "\nALTER TABLE " + self.tabela +
                    " ADD " + coluna + '_' + self.fk_ + " varchar(500);"
                    "\nend;")
        return sql


class Excel(Importavel):

    """ docstring """

    def __init__(self, planilha):
        super(Excel, self).__init__()
        self.tabela = None
        self.arquivo = None
        self.planilha = planilha

    def __str__(self):
        """ Sobreescrita do método de classe str, que é a forma que a o objeto
        é descrito como string.

        Returns:
            Nome da classe.
        """
        return str(self.__class__.__name__)+' '+self.planilha

    def select(self, offset, limit):
        """ docstring """
        work_book = load_workbook(
            filename='../data/' + self.arquivo, read_only=True)
        work_sheet = work_book[self.planilha]
        return work_sheet.rows

    def count(self):
        """ docstring """
        work_book = load_workbook(
            filename='../data/' + self.arquivo, read_only=True)
        work_sheet = work_book[self.planilha]
        return len([row for row in work_sheet.rows])
